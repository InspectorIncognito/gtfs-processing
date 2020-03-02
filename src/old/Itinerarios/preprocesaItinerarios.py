from openpyxl import Workbook
from openpyxl import load_workbook
import csv

wb = load_workbook('PO 2016(01Jul al 31Dic) UN1 - Anexo 5.xlsx')
ws = wb.get_sheet_by_name('01072016 al 31122016')

#last_row = ws.get_highest_row()

#print ws.get_highest_row()
#print last_row
#print ws.last_row()

cell_range = ws['A7':'P630']
tuple(ws.iter_rows('A7:P630'))

for row in ws.iter_rows('A7:P630'):
	for cell in row:
		print cell

print wb.get_sheet_names()