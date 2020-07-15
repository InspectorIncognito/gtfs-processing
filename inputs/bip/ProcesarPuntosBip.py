
# coding: utf-8

# In[2]:


import urllib
import os
import requests
import re
import csv
import hashlib

from openpyxl import load_workbook
from lxml import html


project_path = 'C:\\workspace\\GTFSProcessing\\inputs\\bip'


# # Descargar archivos bip

# In[3]:


# tuple (filename, file_url, first_cell_with_data)
files = [
    ('retail.xlsx', 'http://datos.gob.cl/dataset/33353'),
    ('bip_point.xlsx', 'http://datos.gob.cl/dataset/28198'),
    ('bip_center_high_level.xlsx', 'http://datos.gob.cl/dataset/28192'),
    ('bip_center_standard_level.xlsx', 'http://datos.gob.cl/dataset/28194'),
    ('subway.xlsx', 'http://datos.gob.cl/dataset/33355'),
]

for filename, url in files:
    page = requests.get(url)
    tree = html.fromstring(page.content)
    file_url = tree.xpath('//*[@id="dataset-resources"]/ul/li/div/ul/li[2]/a/@href')[0]
    urllib.request.urlretrieve(file_url, os.path.join(project_path, filename))
    print('file {0} downloaded!'.format(filename))


# In[4]:


# clases para procesar los archivos 
class FileBaseManager(object):
    
    def __init__(self):
        self.empty_schedule_counter = 0
        
    def format_schedule(self, value):
        schedule = []
        previous_value = value
        
        if value is None:
            self.empty_schedule_counter += 1
            return ''

        # delete multiple spaces
        value = ' '.join(value.split())
        
        value = value.replace(' :', ':')
        value = value.replace('Vie:', 'Vie').replace('Vier:', 'Vie')
        value = value.replace('Sab:', 'Sab').replace('Sáb:', 'Sab').replace('Sábados', 'Sab')
        value = value.replace('Lunes a Viernes', 'Lun-Vie').replace('Lun a Vie', 'Lun-Vie')
        value = value.replace('Lun a Dom', 'Lun-Dom').replace('Lun a Sab', 'Lun-Sab')
        value = value.replace(' - ', ' a ').replace('Dom:', 'Dom')

        patterns_1 = [
            'Lun-Vie \d{2}:\d{2} *a *\d{2}:\d{2} \/ *\d{2}:\d{2} a \d{2}:\d{2}',
            'Lun-Vie \d{1,2}:\d{2} *a *\d{2}:\d{2}',
            'Lun-Dom \d{2}:\d{2} *a *\d{2}:\d{2}',
            'Lun-Sab \d{2}:\d{2} *a *\d{2}:\d{2} \/ *\d{2}:\d{2} a \d{2}:\d{2}',
            'Lun-Sab \d{2}:\d{2} *a *\d{2}:\d{2}',
        ]
        patterns_2 = [
            ' Sab \d{1,2}:\d{2} a \d{2}:\d{2}',
            'Sab-Dom \d{2}:\d{2} a \d{2}:\d{2}'
        ]
        patterns_3 = [
            ' Dom \d{2}:\d{2} a \d{2}:\d{2}',
            ' Dom y Festivos \d{2}:\d{2} a \d{2}:\d{2}',
            'Festivos \d{2}:\d{2} a \d{2}:\d{2}',
        ]

        for patterns in [patterns_1, patterns_2, patterns_3]:
            for pattern in patterns:
                result = list(map(lambda x: x.strip(), re.findall(pattern, value)))
                if len(result) != 0:
                    schedule += result
                    break

        value = ','.join(schedule)
        return previous_value if value == '' else value
        
    def get_data(self, sheet):
        data = []
        for row in sheet.iter_rows(min_row=self.first_row, max_col=sheet.max_column, max_row=sheet.max_row):
            row_data = [row[col_index].value if isinstance(col_index, int) else col_index() for col_index in self.columns]
            schedule_index = self.columns.index(self.schedule_column)
            row_data[schedule_index] = self.format_schedule(str(row_data[schedule_index]))
            row_data.append(self.services)
            data.append(row_data)
        
        return data

class RetailFileManager(FileBaseManager):
    
    def __init__(self):
        super().__init__()
        self.services = 'Carga de tarjetas,Consulta de saldo, Activación de carga remota'
        # column starts from zero
        self.columns = [1, 2, 3, 5, 8, 9]
        self.schedule_column = 5
        self.first_row = 12

class BipCenterHighLevelFileManager(FileBaseManager):
    
    def __init__(self):
        super().__init__()
        self.services = 'Venta de tarjetas,Carga de tarjetas,Consulta de saldo, Activación de carga remota,Reemplazo de tarjeta,Recuperación de saldo de tarjetas dañadas'
        # column starts from zero
        # we use column index of excel file or lambda function to print text directly
        self.columns = [2, lambda: "Centro bip!", 3, 5, 8, 9]
        self.schedule_column = 5
        self.first_row = 16

class BipCenterStandardLevelFileManager(FileBaseManager):
    
    def __init__(self):
        super().__init__()
        self.services = 'Venta de tarjetas,Carga de tarjetas,Consulta de saldo, Activación de carga remota'
        # column starts from zero
        # we use column zero because this file does not have name column
        self.columns = [2, lambda: "Centro bip!", 3, 5, 8, 9]
        self.schedule_column = 5
        self.first_row = 13


# # Pruebas para estandarizar formatos de horario

# In[5]:


# testing
tests = [
# retail
    ('Lun-Vie 08:30 a 22:00 Sab 08:30 a 22:00 Dom 08:30 a 22:00', 'Lun-Vie 08:30 a 22:00,Sab 08:30 a 22:00,Dom 08:30 a 22:00'),
    ('Lun-Dom 09:00 a 21:00', 'Lun-Dom 09:00 a 21:00'),
    ('Lun-Dom 08:30 a 22:00', 'Lun-Dom 08:30 a 22:00'),
# bip points
    ('Lun a Vie 8:00 a 19:00 Sab 9:00 a 17:00', 'Lun-Vie 8:00 a 19:00,Sab 9:00 a 17:00'),
    ('Lun a Vie: 09:00 a 17:30', 'Lun-Vie 09:00 a 17:30'),
    ('Lun a Vie: 09:00 a 19:00 - Sab: 11:00 a 14:00', 'Lun-Vie 09:00 a 19:00,Sab 11:00 a 14:00'),
    ('Lun a Vie: 09:00 a 19:00 / 15:00 a 19:00 - Sab: 10:00 a 14:00', 'Lun-Vie 09:00 a 19:00 / 15:00 a 19:00,Sab 10:00 a 14:00'),
    ('Lun a Vie: 10:00 a 19:00', 'Lun-Vie 10:00 a 19:00'),
    ('Lun a Dom: 10:00 a 20:00', 'Lun-Dom 10:00 a 20:00'),
    ('Lun a Vie: 09:00 a 21:00 - Sab: 09:00 a 21:00', 'Lun-Vie 09:00 a 21:00,Sab 09:00 a 21:00'),
    ('Lun a Vie: 09:00 a 20:00 - Sab-Dom: 10:30 a 20:00', 'Lun-Vie 09:00 a 20:00,Sab-Dom 10:30 a 20:00'),
    ('Lun a Sab: 10:00 a 21:00 - Dom: 11:00 a 21:00', 'Lun-Sab 10:00 a 21:00,Dom 11:00 a 21:00'),
    ('Lun a Vie: 10:00 a 14:00 / 15:00 a 19:00 - Sab: 10:00 a 14:00', 'Lun-Vie 10:00 a 14:00 / 15:00 a 19:00,Sab 10:00 a 14:00'),
    ('Lun a Sab: 08:00 a 22:00 - Dom: 09:00 a 22:00', 'Lun-Sab 08:00 a 22:00,Dom 09:00 a 22:00'),
    ('Lun a Sab: 09:00 a 21:00', 'Lun-Sab 09:00 a 21:00'),
    ('Lun a Vie: 11:00 - 15:00 / 16:00 - 20:00 - Sab: 11:00 a 15:00', 'Lun-Vie 11:00 a 15:00 / 16:00 a 20:00,Sab 11:00 a 15:00'),
    ('Lun a Dom: 08:30 a 21:30', 'Lun-Dom 08:30 a 21:30'),
    ('Lun a Sab: 09:00 a 15:00 /15:30 a 21:00', 'Lun-Sab 09:00 a 15:00 /15:30 a 21:00'),
    ('Lun a Dom: 09:00 a 21:00', 'Lun-Dom 09:00 a 21:00'),
    ('Lun a Vier: 10:00 a 14:00 / 15:00 a 18:30 / Sáb: 10:00 a 14:00', 'Lun-Vie 10:00 a 14:00 / 15:00 a 18:30,Sab 10:00 a 14:00'),
    ('Lun a Vie : 11:00 a 15:00 /16:00 a 19:00 - Sab: 10:00 a 15:00', 'Lun-Vie 11:00 a 15:00 /16:00 a 19:00,Sab 10:00 a 15:00'),
    ('Lun a Vie: 11:00 - 15:00 / 16:00 - 19:00 - Sab: 10:00 a 14:00', 'Lun-Vie 11:00 a 15:00 / 16:00 a 19:00,Sab 10:00 a 14:00'),
    ('Lun a Vie: 09:00 a 20:00 - Sab: 09:00 a 14:00', 'Lun-Vie 09:00 a 20:00,Sab 09:00 a 14:00'),
    ('Lun a Vie: 10:00 a 19:00 - Sab: 10:00 a 14:00', 'Lun-Vie 10:00 a 19:00,Sab 10:00 a 14:00'),
    ('Lun a Sab: 08:30 a 22:00- Dom: 09:00 a 22:00', 'Lun-Sab 08:30 a 22:00,Dom 09:00 a 22:00'),
    ('Lun a Vie: 09:00 a 21:00 - Sab: 09:00 a 14:00', 'Lun-Vie 09:00 a 21:00,Sab 09:00 a 14:00'),
    ('Lun a Vie: 10:00 a 14:30 / Sab: 15:30 a 18:30', 'Lun-Vie 10:00 a 14:30,Sab 15:30 a 18:30'),
    ('Lun a Vie: 09:00 a 20:00 - Sab: 10:00 a 14:00', 'Lun-Vie 09:00 a 20:00,Sab 10:00 a 14:00'),
    ('Lun a Vie: 10:00 a 20:00 - Sab: 10:00 a 14:00', 'Lun-Vie 10:00 a 20:00,Sab 10:00 a 14:00'),
    ('Lun a Dom: 09:00 a 20:00', 'Lun-Dom 09:00 a 20:00'),
    ('Lunes a Viernes 11:00 a 14:30 / 15:30 a 20:00', 'Lun-Vie 11:00 a 14:30 / 15:30 a 20:00'),
    ('Lunes a Viernes 10:00 a 14:00 / 15:00 a 18:30; Sábados 10:00 a 14:00', 'Lun-Vie 10:00 a 14:00 / 15:00 a 18:30,Sab 10:00 a 14:00'),
    ('Lunes a Viernes 09:00 a 14:30 / 15:30 a 18:30', 'Lun-Vie 09:00 a 14:30 / 15:30 a 18:30'),
    ('Lun a Vie 07:00 a 22:00 Sab 08:00 a 22:00 Dom y Festivos 09:00 a 22:00', 'Lun-Vie 07:00 a 22:00,Sab 08:00 a 22:00,Dom y Festivos 09:00 a 22:00'),
    ('Lun a Sab 08:00 a 21:00 Dom y Festivos 09:00 a 21:00', 'Lun-Sab 08:00 a 21:00,Dom y Festivos 09:00 a 21:00'),
    ('Lun a Sab 08:00 a 21:00 Dom y Festivos 09:00 a 14:00', 'Lun-Sab 08:00 a 21:00,Dom y Festivos 09:00 a 14:00'),
    ('Lun a Sab 08:30 a 21:30 Dom y Festivos 09:00 a 21:00', 'Lun-Sab 08:30 a 21:30,Dom y Festivos 09:00 a 21:00'),
    ('Lun a Sab 08:30 a 21:00 Dom y Festivos 09:00 a 14:00', 'Lun-Sab 08:30 a 21:00,Dom y Festivos 09:00 a 14:00'),
    ('Lun a Dom 09:00 a 22:00 Festivos 09:00 a 22:00', 'Lun-Dom 09:00 a 22:00,Festivos 09:00 a 22:00'),
    ('Lun a Vie 05:30 a 23:00 Sab 06:30 a 23:00 Dom y Festivos 08:00 a 23:00', 'Lun-Vie 05:30 a 23:00,Sab 06:30 a 23:00,Dom y Festivos 08:00 a 23:00'),
    ('Lun a Vie 06:00 a 23:00 Sab 07:00 a 23:00 Dom y Festivos 08:00 a 23:00', 'Lun-Vie 06:00 a 23:00,Sab 07:00 a 23:00,Dom y Festivos 08:00 a 23:00'),
    ('Lun a Dom : 09:00 a 21:00', 'Lun-Dom 09:00 a 21:00'),
    ('asdfsdfsdf', 'asdfsdfsdf')
]
analizer = FileBaseManager()
for index, (test_value, expected_value) in enumerate(tests):
    value = analizer.format_schedule(test_value)
    assert value == expected_value, 'Error in index {0}: \n{1}\n{2}'.format(index, value, expected_value)


# # Leer datos de cada archivo

# In[6]:


output_data = []
empty_schedule_counter = 0

for filename, _ in files:
    wb = load_workbook(filename=os.path.join(project_path, filename), read_only=True)
    sheet= wb[wb.sheetnames[0]]

    print (filename)

    if filename in ['retail.xlsx', 'bip_point.xlsx']:
        file_manager = RetailFileManager()
    elif filename in ['bip_center_high_level.xlsx']:
        file_manager = BipCenterHighLevelFileManager()
    elif filename in ['bip_center_standard_level.xlsx']:
        file_manager = BipCenterStandardLevelFileManager()
    else:
        continue
    
    output_data += file_manager.get_data(sheet)
    empty_schedule_counter += file_manager.empty_schedule_counter

print(empty_schedule_counter)
print(len(output_data))


# # Guardar datos en CSV

# In[7]:


output_filename = 'PhoneFairPoints.csv'
output_path = os.path.join(project_path, output_filename)
header = ['id', 'entity', 'name', 'address', 'schedule', 'longitude', 'latitude', 'services']

with open(output_path, 'w',newline='') as csvfile:
    spamwriter = csv.writer(csvfile, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)
    spamwriter.writerow(header)
    for index, row in enumerate(output_data):
        spamwriter.writerow([index + 1] + row)


# # Revisar si cambió

# In[8]:


checksum_filename = 'checksum'
checksum_file_path = os.path.join(project_path, checksum_filename)

current_checksum = hashlib.md5(open(output_path, 'rb').read()).hexdigest()

if os.path.isfile(checksum_file_path):
    with open(checksum_file_path) as checksum_file_obj:
        previous_checksum = checksum_file_obj.read()
    if current_checksum == previous_checksum:
        print('The file has not changed')
    else:
        print('File was updated')
        with open(checksum_file_path, 'w') as checksum_file_obj:
            checksum_file_obj.write(current_checksum)
else:
    print('There is not a checksum file, so i will create it')
    with open(checksum_file_path, 'w') as checksum_file_obj:
        checksum_file_obj.write(current_checksum)

